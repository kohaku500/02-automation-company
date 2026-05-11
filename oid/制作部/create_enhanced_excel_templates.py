#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excelテンプレートを完全刷新
AI中心の実装内容に更新・Day 1-30の詳細ステップを含める
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class EnhancedExcelTemplateGenerator:
    def __init__(self):
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.base_path = base_path

        self.colors = {
            'header': PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid"),
            'header2': PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid"),
            'bg': PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
            'success': PatternFill(start_color="C8F5C8", end_color="C8F5C8", fill_type="solid"),
        }
        self.font_header = Font(bold=True, color="FFFFFF", size=11)
        self.font_normal = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def create_sales_excel(self):
        """営業自動化テンプレート"""
        wb = Workbook()

        # シート1: メール生成ガイド
        ws = wb.active
        ws.title = "メール生成"
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 35

        # ヘッダー
        ws['A1'] = "AI時代の個人スキル販売術"
        ws['A1'].font = Font(bold=True, size=14, color="667EEA")
        ws.merge_cells('A1:C1')

        ws['A2'] = "ChatGPT営業メール自動化 Day 1-30実装ガイド"
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:C2')

        # Day別実装ガイド
        row = 4
        days_content = [
            ("Day 1-2", "基礎習得", "ChatGPT/Gemini登録・初回メール生成"),
            ("Day 3-7", "最適化", "営業メール生成プロンプト3パターン実験"),
            ("Day 8-14", "実運用", "毎日30-50通のメール配信ルーチン"),
            ("Day 15-21", "分析・改善", "返信データ分析→メール改善"),
            ("Day 22-28", "自動化", "Excel マクロ + Gemini自動分析"),
            ("Day 29-30", "評価", "月間KPI評価・翌月改善計画"),
        ]

        for day, phase, content in days_content:
            ws[f'A{row}'] = day
            ws[f'B{row}'] = phase
            ws[f'C{row}'] = content

            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.font = self.font_normal
                cell.fill = self.colors['bg']
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            ws.row_dimensions[row].height = 30
            row += 1

        # ChatGPTプロンプト集
        row += 2
        ws[f'A{row}'] = "【ChatGPTプロンプト集】"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="667EEA")
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        prompts = [
            ("初回接触メール", "新規見込客向けメール生成（Day 3）"),
            ("提案メール", "ヒアリング後の提案メール生成（Day 4-5）"),
            ("返信分析", "Geminiで顧客ニーズ分析（Day 12-14）"),
            ("改善版メール", "反応が高いメール構成を学習（Day 18-20）"),
        ]

        for prompt_name, description in prompts:
            ws[f'A{row}'] = prompt_name
            ws[f'B{row}'] = description
            ws[f'A{row}'].font = self.font_header
            ws[f'A{row}'].fill = self.colors['header']
            ws[f'B{row}'].fill = self.colors['header']
            ws.merge_cells(f'B{row}:C{row}')
            row += 1

            # プロンプト内容
            prompts_detail = {
                "初回接触メール": '新規営業メール。相手企業:[業種]。課題:[課題]。提案:[ソリューション]。ROI:[削減額]。件名3案+本文。',
                "提案メール": 'ヒアリング後の提案メール。背景:課題確認済み。提案:[内容]。形式:謝礼→確認→提案→価格→次のステップ。',
                "返信分析": 'メール分析。観点1:顧客のニーズ把握 2:説得力 3:返信を促す工夫。',
                "改善版メール": '改善版メール。指摘:[Geminiからの指摘]。これらを反映して改善版を作成。',
            }

            ws[f'A{row}'] = prompts_detail[prompt_name]
            ws[f'A{row}'].font = self.font_normal
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{row}:C{row}')
            ws.row_dimensions[row].height = 40
            row += 2

        # シート2: KPI分析
        ws2 = wb.create_sheet("KPI分析")
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15

        # KPI表
        kpi_header = ['KPI項目', '実績', '目標']
        for col, header in enumerate(kpi_header, 1):
            cell = ws2.cell(row=1, column=col)
            cell.value = header
            cell.font = self.font_header
            cell.fill = self.colors['header']
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        kpis = [
            ("送信メール数", "=COUNTA(C2:C31)", "1000"),
            ("返信メール数", "", "280"),
            ("返信率", "=IF(B2>0,B3/B2,0)", "28%"),
            ("成約見込み数", "", "42"),
            ("月間営業時間削減", "=450*0.80", "360時間"),
            ("月間増収見込み", "=B4*3000000", "¥1,260万"),
        ]

        for idx, (item, formula, target) in enumerate(kpis, 2):
            ws2[f'A{idx}'] = item
            ws2[f'B{idx}'] = formula
            ws2[f'C{idx}'] = target

            for col in ['A', 'B', 'C']:
                cell = ws2[f'{col}{idx}']
                cell.font = self.font_normal
                cell.border = self.border
                cell.alignment = Alignment(horizontal='right' if col != 'A' else 'left')

        # シート3: 日次実行記録
        ws3 = wb.create_sheet("日次記録")
        ws3.column_dimensions['A'].width = 12
        ws3.column_dimensions['B'].width = 15
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 20

        headers = ['日付', '送信数', '返信数', '返信メール内容']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col)
            cell.value = header
            cell.font = self.font_header
            cell.fill = self.colors['header']
            cell.border = self.border

        # 30日分のテンプレート行
        for day in range(1, 31):
            row = day + 1
            ws3[f'A{row}'] = f"Day {day}"
            ws3[f'B{row}'] = ""  # 送信数を入力
            ws3[f'C{row}'] = ""  # 返信数を入力
            ws3[f'D{row}'] = ""  # 返信内容メモ

        return wb

    def create_sns_excel(self):
        """SNS自動化テンプレート"""
        wb = Workbook()

        ws = wb.active
        ws.title = "投稿スケジュール"
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

        # ヘッダー
        ws['A1'] = "SNS運用自動化キット"
        ws['A1'].font = Font(bold=True, size=14, color="764BA2")
        ws.merge_cells('A1:E1')

        ws['A2'] = "ChatGPT投稿文自動生成 Day 1-30実装ガイド"
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:E2')

        # Day別パターン
        row = 4
        days = [
            ("Day 1-2", "基礎習得", "ChatGPT投稿自動生成の基本習得"),
            ("Day 3-7", "パターン確立", "曜日別投稿パターン7種類を確立"),
            ("Day 8-14", "実運用", "毎日1投稿（曜日別パターン実施）"),
            ("Day 15-21", "分析・改善", "エンゲージ率を分析・改善"),
            ("Day 22-28", "自動化", "投稿スケジュール完全自動化"),
            ("Day 29-30", "評価", "月間KPI評価・翌月計画"),
        ]

        for day, phase, content in days:
            ws[f'A{row}'] = day
            ws[f'B{row}'] = phase
            ws[f'C{row}'] = content

            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.font = self.font_normal
                cell.fill = self.colors['bg']
                cell.border = self.border

            row += 1

        # 曜日別投稿パターン
        row += 2
        ws[f'A{row}'] = "【曜日別投稿パターン】"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="764BA2")
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        day_patterns = [
            ("月", "トレンド", "新情報・業界ニュース"),
            ("火", "How-To", "具体的な実装方法"),
            ("水", "ビジュアル", "画像+短い解説"),
            ("木", "ストーリー", "個人的な体験・成長"),
            ("金", "励まし", "モチベーション・応援"),
            ("土", "知識", "データ・学習情報"),
            ("日", "計画", "目標設定・来週準備"),
        ]

        for day, pattern, content in day_patterns:
            ws[f'A{row}'] = day
            ws[f'B{row}'] = pattern
            ws[f'C{row}'] = content

            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.font = self.font_normal
                if row % 2 == 0:
                    cell.fill = self.colors['bg']
                cell.border = self.border

            row += 1

        # シート2: KPI分析
        ws2 = wb.create_sheet("KPI分析")
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15

        kpi_header = ['KPI項目', '実績', '目標']
        for col, header in enumerate(kpi_header, 1):
            cell = ws2.cell(row=1, column=col)
            cell.value = header
            cell.font = self.font_header
            cell.fill = self.colors['header2']
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        kpis = [
            ("月間投稿数", "=COUNTA(B2:B31)", "30"),
            ("フォロワー増加", "=B2-A2", "100"),
            ("いいね合計", "", "3,200"),
            ("エンゲージ率", "=IF(C2>0,B3/C2,0)", "6.4%"),
            ("営業接触件数", "", "15"),
            ("月間運用時間削減", "=35*30", "17.5時間"),
        ]

        for idx, (item, formula, target) in enumerate(kpis, 2):
            ws2[f'A{idx}'] = item
            ws2[f'B{idx}'] = formula if formula else ""
            ws2[f'C{idx}'] = target

            for col in ['A', 'B', 'C']:
                cell = ws2[f'{col}{idx}']
                cell.font = self.font_normal
                cell.border = self.border
                cell.alignment = Alignment(horizontal='right' if col != 'A' else 'left')

        # シート3: 日次記録
        ws3 = wb.create_sheet("日次記録")
        ws3.column_dimensions['A'].width = 12
        ws3.column_dimensions['B'].width = 15
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 15
        ws3.column_dimensions['E'].width = 20

        headers = ['日付', '曜日別パターン', 'いいね数', 'RT数', 'メモ']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col)
            cell.value = header
            cell.font = self.font_header
            cell.fill = self.colors['header2']
            cell.border = self.border

        # 30日分のテンプレート
        day_names = ["月", "火", "水", "木", "金", "土", "日"]
        for day in range(1, 31):
            row = day + 1
            ws3[f'A{row}'] = f"Day {day}"
            ws3[f'B{row}'] = day_names[(day-1) % 7]
            ws3[f'C{row}'] = ""  # いいね数
            ws3[f'D{row}'] = ""  # RT数
            ws3[f'E{row}'] = ""  # メモ

        return wb

    def create_ai_guide_excel(self):
        """初心者向けAIガイド"""
        wb = Workbook()

        ws = wb.active
        ws.title = "実装ロードマップ"
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 10

        # ヘッダー
        ws['A1'] = "初心者向けAI活用ガイド"
        ws['A1'].font = Font(bold=True, size=14, color="4CAF50")
        ws.merge_cells('A1:D1')

        ws['A2'] = "ChatGPT/Gemini 30日実装マスタープログラム"
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:D2')

        # テーブルヘッダー
        headers = ['フェーズ', '期間', '実装内容', '完了']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = self.font_header
            cell.fill = self.colors['success']
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        # フェーズデータ
        phases = [
            ("ChatGPT基本習得", "Day 1-5", "登録→基本操作→実務応用"),
            ("Gemini活用習得", "Day 6-10", "登録→分析→改善"),
            ("業務別AI活用", "Day 11-15", "メール・提案文・データ分析"),
            ("自動化・効率化", "Day 16-20", "Excel連携→マクロ→自動実行"),
            ("スキルアップ", "Day 21-25", "プロンプトエンジニアリング学習"),
            ("継続改善", "Day 26-30", "月間評価→翌月計画→成長サイクル"),
        ]

        row = 5
        for phase, period, content in phases:
            ws[f'A{row}'] = phase
            ws[f'B{row}'] = period
            ws[f'C{row}'] = content
            ws[f'D{row}'] = "☐"

            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.font = self.font_normal
                cell.border = self.border
                if row % 2 == 0:
                    cell.fill = self.colors['bg']

            row += 1

        # 期待値
        row += 2
        ws[f'A{row}'] = "期待される成果"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="4CAF50")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        expectations = [
            "日々の手作業が3-4時間→1時間に削減（70%削減）",
            "月間削減時間: 60-80時間",
            "同じ時間でできる仕事量: 3倍以上に増加",
            "新規プロジェクト開始に充当可能な時間確保",
            "生産性向上による月間追加売上見込み: ¥500万以上",
        ]

        for expectation in expectations:
            ws[f'A{row}'] = expectation
            ws[f'A{row}'].font = self.font_normal
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

        return wb

    def run(self):
        print("[Claude Design] Excelテンプレートを完全刷新中...\n")

        products = [
            ("AI時代の個人スキル販売術", self.create_sales_excel),
            ("SNS運用自動化キット", self.create_sns_excel),
            ("初心者向けAI活用ガイド", self.create_ai_guide_excel),
        ]

        for name, gen_func in products:
            print(f"【{name}】")
            product_dir = Path(os.path.join(self.base_path, f"生成物・商品/素材/{name}"))

            # 古いファイルを削除
            for pattern in [".xlsx", ".xlsm"]:
                old_file = product_dir / f"{name}_テンプレート{pattern}"
                if old_file.exists():
                    old_file.unlink()

            excel_file = product_dir / f"{name}_テンプレート.xlsx"

            wb = gen_func()
            wb.save(str(excel_file))

            print(f"  ✅ {name}_テンプレート.xlsx")
            print(f"     シート数: {len(wb.sheetnames)} | ファイルサイズ: {os.path.getsize(excel_file) / 1024:.1f}KB\n")

        print("✅ Excelテンプレート完全刷新完了")

if __name__ == "__main__":
    generator = EnhancedExcelTemplateGenerator()
    generator.run()
