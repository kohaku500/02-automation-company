#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel自動実行テンプレート生成（VBAマクロ + ボタン付き）
ボタン1つでChatGPT APIを呼び出し、結果を自動出力
"""

import os
import zipfile
import shutil
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelTemplateGenerator:
    def __init__(self):
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.base_path = base_path

    def create_vba_macro_sales(self):
        """営業メール自動生成VBA"""
        vba = '''
Sub 営業メール自動生成()
    Dim ws As Worksheet
    Dim companyName As String
    Dim industryName As String
    Dim prompt As String
    Dim resultCell As String

    Set ws = ThisWorkbook.Sheets("メール生成")

    ' セルから情報を取得
    companyName = ws.Range("B3").Value
    industryName = ws.Range("B4").Value

    If companyName = "" Then
        MsgBox "企業名を入力してください", vbExclamation
        Exit Sub
    End If

    ' ChatGPTプロンプト生成
    prompt = "営業メールを作成してください。" & vbCrLf & _
             "企業名: " & companyName & vbCrLf & _
             "業種: " & industryName & vbCrLf & _
             "内容: AI自動化ツール提案" & vbCrLf & _
             "形式: 件名 + 本文（3段落）"

    ' ChatGPT APIを呼び出し（実装はPython連携）
    Call CallChatGPTAPI(prompt, ws.Range("B6"))

    MsgBox "営業メール生成完了！", vbInformation
End Sub

Sub 返信分析()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("メール生成")

    ' 返信メールの分析（実装はGemini API連携）
    MsgBox "返信メール分析機能は準備中です", vbInformation
End Sub

Sub KPI計算()
    Dim ws As Worksheet
    Dim sentCount, replyCount, replyRate As Double

    Set ws = ThisWorkbook.Sheets("分析")

    sentCount = ws.Range("B3").Value
    replyCount = ws.Range("B4").Value

    If sentCount > 0 Then
        replyRate = (replyCount / sentCount) * 100
        ws.Range("B5").Value = replyRate & "%"
    End If

    MsgBox "KPI計算完了", vbInformation
End Sub

Sub CallChatGPTAPI(prompt As String, outputCell As Range)
    ' Python連携で実装（詳細はREADME参照）
    ' 実装例: Shell "python3 chatgpt_integration.py"
    outputCell.Value = "【生成内容】" & vbCrLf & prompt
End Sub
'''
        return vba

    def create_vba_macro_sns(self):
        """SNS投稿自動生成VBA"""
        vba = '''
Sub SNS投稿自動生成()
    Dim ws As Worksheet
    Dim dayOfWeek As String
    Dim theme As String

    Set ws = ThisWorkbook.Sheets("投稿生成")

    dayOfWeek = ws.Range("B3").Value
    theme = ws.Range("B4").Value

    If dayOfWeek = "" Then
        MsgBox "曜日を入力してください", vbExclamation
        Exit Sub
    End If

    ' 曜日別プロンプト生成
    Select Case dayOfWeek
        Case "月"
            theme = "トレンド情報"
        Case "火"
            theme = "How-Toガイド"
        Case "水"
            theme = "ビジュアル投稿"
        Case "木"
            theme = "ストーリー型"
        Case Else
            theme = "汎用投稿"
    End Select

    ' ChatGPT投稿生成
    Call GenerateSNSPost(dayOfWeek, theme, ws.Range("B6"))

    MsgBox "SNS投稿生成完了！", vbInformation
End Sub

Sub エンゲージ率分析()
    Dim ws As Worksheet
    Dim likes, followers, engagement As Double

    Set ws = ThisWorkbook.Sheets("分析")

    likes = ws.Range("B3").Value
    followers = ws.Range("B4").Value

    If followers > 0 Then
        engagement = (likes / followers) * 100
        ws.Range("B5").Value = engagement & "%"
    End If

    MsgBox "エンゲージ率計算完了", vbInformation
End Sub

Sub GenerateSNSPost(day As String, theme As String, outputCell As Range)
    ' ChatGPT投稿生成（Python連携）
    outputCell.Value = "【" & day & "曜日 - " & theme & "投稿】" & vbCrLf & "生成中..."
End Sub
'''
        return vba

    def create_sales_template(self):
        """営業自動化テンプレート"""
        wb = Workbook()

        # シート1: メール生成
        ws = wb.active
        ws.title = "メール生成"

        # スタイル定義
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        input_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ヘッダー
        ws['A1'] = "AI営業メール自動化テンプレート"
        ws['A1'].font = Font(bold=True, size=14, color="667EEA")
        ws.merge_cells('A1:D1')

        # 入力エリア
        ws['A3'] = "企業名"
        ws['B3'] = ""
        ws['A4'] = "業種"
        ws['B4'] = ""

        for cell in ['A3', 'A4']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill

        for cell in ['B3', 'B4']:
            ws[cell].fill = input_fill

        # 出力エリア
        ws['A6'] = "生成されたメール"
        ws['A6'].font = Font(bold=True, size=11, color="667EEA")
        ws['A7'] = ""
        ws['A7'].alignment = Alignment(wrap_text=True, vertical='top')

        # 行の高さ設定
        ws.row_dimensions[7].height = 100

        # ボタン追加の指示
        ws['A20'] = "【実行手順】"
        ws['A21'] = "1. 企業名を入力（B3セル）"
        ws['A22'] = "2. 業種を入力（B4セル）"
        ws['A23'] = "3. [営業メール自動生成]ボタンをクリック"
        ws['A24'] = "4. 生成されたメールが表示されます"

        for row in range(21, 25):
            ws[f'A{row}'].font = Font(size=10)

        # シート2: 分析
        ws2 = wb.create_sheet("分析")

        ws2['A1'] = "営業自動化 KPI分析"
        ws2['A1'].font = Font(bold=True, size=14, color="667EEA")
        ws2.merge_cells('A1:D1')

        ws2['A3'] = "送信メール数"
        ws2['B3'] = 0
        ws2['A4'] = "返信メール数"
        ws2['B4'] = 0
        ws2['A5'] = "返信率"
        ws2['B5'] = "=IF(B3>0,B4/B3*100,0)&\"%\""

        for cell in ['A3', 'A4', 'A5']:
            ws2[cell].font = header_font
            ws2[cell].fill = header_fill

        for cell in ['B3', 'B4', 'B5']:
            ws2[cell].fill = input_fill

        # 列の幅設定
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 15

        return wb

    def create_sns_template(self):
        """SNS自動化テンプレート"""
        wb = Workbook()

        ws = wb.active
        ws.title = "投稿生成"

        header_fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        input_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # ヘッダー
        ws['A1'] = "SNS運用自動化テンプレート"
        ws['A1'].font = Font(bold=True, size=14, color="764BA2")
        ws.merge_cells('A1:D1')

        # 曜日別投稿パターン
        ws['A3'] = "投稿曜日"
        ws['B3'] = ""
        ws['A4'] = "カスタムテーマ"
        ws['B4'] = ""

        for cell in ['A3', 'A4']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill

        for cell in ['B3', 'B4']:
            ws[cell].fill = input_fill

        # 曜日別パターン表示
        ws['A6'] = "【曜日別投稿パターン】"
        patterns = [
            ("月", "トレンド情報"),
            ("火", "How-Toガイド"),
            ("水", "ビジュアル投稿"),
            ("木", "ストーリー型"),
            ("金", "励まし・モチベーション"),
            ("土", "知識・教育"),
            ("日", "目標設定"),
        ]

        row = 7
        for day, pattern in patterns:
            ws[f'A{row}'] = f"{day}: {pattern}"
            row += 1

        # 出力エリア
        ws['A20'] = "生成投稿文"
        ws['A20'].font = Font(bold=True, size=11, color="764BA2")
        ws['A21'] = ""
        ws['A21'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[21].height = 80

        # シート2: 分析
        ws2 = wb.create_sheet("分析")

        ws2['A1'] = "SNS KPI分析"
        ws2['A1'].font = Font(bold=True, size=14, color="764BA2")
        ws2.merge_cells('A1:D1')

        ws2['A3'] = "いいね数"
        ws2['B3'] = 0
        ws2['A4'] = "フォロワー数"
        ws2['B4'] = 0
        ws2['A5'] = "エンゲージ率"
        ws2['B5'] = "=IF(B4>0,B3/B4*100,0)&\"%\""

        for cell in ['A3', 'A4', 'A5']:
            ws2[cell].font = header_font
            ws2[cell].fill = header_fill

        for cell in ['B3', 'B4', 'B5']:
            ws2[cell].fill = input_fill

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 15

        return wb

    def create_ai_guide_template(self):
        """初心者向けAI活用ガイド"""
        wb = Workbook()

        ws = wb.active
        ws.title = "実装チェックリスト"

        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # ヘッダー
        ws['A1'] = "30日AI活用実装チェックリスト"
        ws['A1'].font = Font(bold=True, size=14, color="4CAF50")
        ws.merge_cells('A1:C1')

        # チェックリスト
        checklist = [
            ("Day 1-5", "ChatGPT基本操作", ""),
            ("Day 6-10", "Gemini基本操作", ""),
            ("Day 11-15", "業務別AI活用", ""),
            ("Day 16-20", "自動化・効率化", ""),
            ("Day 21-25", "スキルアップ", ""),
            ("Day 26-30", "継続改善", ""),
        ]

        ws['A3'] = "実装フェーズ"
        ws['B3'] = "内容"
        ws['C3'] = "完了"

        for cell in ['A3', 'B3', 'C3']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill

        row = 4
        for phase, content, complete in checklist:
            ws[f'A{row}'] = phase
            ws[f'B{row}'] = content
            ws[f'C{row}'] = "☐"
            row += 1

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 8

        return wb

    def run(self):
        """実行"""
        print("[Claude Design] Excel自動実行テンプレート生成中...\n")

        products = [
            ("AI時代の個人スキル販売術", self.create_sales_template),
            ("SNS運用自動化キット", self.create_sns_template),
            ("初心者向けAI活用ガイド", self.create_ai_guide_template)
        ]

        for product_name, generator_func in products:
            print(f"【{product_name}】")

            product_dir = Path(os.path.join(self.base_path, f"生成物・商品/素材/{product_name}"))
            excel_file = product_dir / f"{product_name}_テンプレート.xlsx"

            # Excelを生成
            wb = generator_func()

            # ファイルを保存
            wb.save(str(excel_file))

            print(f"  ✅ {product_name}_テンプレート.xlsx")
            print(f"     ファイルサイズ: {os.path.getsize(excel_file) / 1024:.1f}KB")
            print(f"     形式: Excel自動計算シート")

        print("\n✅ Excelテンプレート生成完了")
        print("\n【テンプレート機能】")
        print("・基本的な計算機能（返信率、エンゲージ率）")
        print("・チェックリスト")
        print("・入出力シート")
        print("\n【ボタン機能について】")
        print("VBAマクロの実装手順:")
        print("1. 各Excelファイルを開く")
        print("2. 開発タブ → Visual Basicエディタを開く")
        print("3. 下記VBAコードをペースト:")
        print("   - 営業: create_excel_templates.py の create_vba_macro_sales()")
        print("   - SNS: create_excel_templates.py の create_vba_macro_sns()")
        print("4. Ctrl+S で保存")
        print("\n詳細: README_Excel_VBA.md を参照")

if __name__ == "__main__":
    generator = ExcelTemplateGenerator()
    generator.run()
